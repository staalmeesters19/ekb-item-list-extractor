"""Fase 1 match-rapport: FCC-13 rijen tegen ProCos export.

Quick & dirty matching:
- Load 86k ProCos artikelen
- Extract 66 rows from FCC-13 PDF
- Manual fabrikant mapping (top NL/DE brands)
- Exact match on (fabrikantcode, normalized Type nr.)
- Report per row: MATCH / NIET UNIEK / NIET GEVONDEN
- Write Excel report

NOT part of the main pipeline — standalone exploration script.
"""

from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml

# Make extractor importable
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "extractor"))
from src.pipeline import run as pipeline_run  # noqa: E402


PROCOS_EXPORT = r"C:\Users\JorisMerkx\Downloads\ProCos-export Artikeldata-excl prijzen.xlsx"
PDF_PATH = r"C:\Users\JorisMerkx\OneDrive - Agyle\Documenten\EKB\test lijst.pdf"
PDF_PAGES = [[8, 9]]
OUT_REPORT = r"C:\Users\JorisMerkx\OneDrive - Agyle\Documenten\EKB\FCC13_match_rapport.xlsx"


# Manual fabrikant mapping — derived from top fabrikantcodes in ProCos export.
# Keys are normalized uppercase tekstnamen die onze extractor uit de PDF haalt.
FAB_MAPPING: dict[str, str] = {
    "SIEMENS": "SIEHAA",
    "SIEMENS AG": "SIEHAA",
    "ROCKWELL": "ROCUIT",
    "ROCKWELL AUTOMATION": "ROCUIT",
    "ABB": "ABBROT1",
    "EATON": "EATVEJ",
    "RITTAL": "RITZEV",
    "PHOENIX": "PHOZEV",
    "PHOENIX CONTACT": "PHOZEV",
    "LAPP": "LAPVEL",
    "WEIDMULLER": "WEIHIL",
    "WEIDMUELLER": "WEIHIL",
    "WEIDMÜLLER": "WEIHIL",
    "SCHNEIDER": "SCHHAA",
    "SCHNEIDER ELECTRIC": "SCHHAA",
    "HARTING": "HARBOS",
    "FESTO": "FESDEL",
    "EAO": "EAODOR",
}


def _normalize_typenr(s: str | None) -> str:
    """Strip punctuation + whitespace + uppercase, mirroring ProCos's
    own match-time normalization (per Werkinstructie)."""
    if not s:
        return ""
    return re.sub(r"[\s\-./()_,]+", "", str(s)).upper()


def load_procos(path: str) -> tuple[dict, dict, int]:
    """Return:
       by_fab_type   : {(fabcode, norm_type): [artikel_records, ...]}
       by_type_only  : {norm_type: [artikel_records, ...]}
       n_rows        : total rows loaded
    """
    print(f"[load] reading {path} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["export"]
    by_fab_type: dict = defaultdict(list)
    by_type_only: dict = defaultdict(list)
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # header
        artikel = row[0]
        fab = row[4]
        type_nr = row[5]
        omsch1 = row[6]
        if not artikel or not type_nr:
            continue
        norm = _normalize_typenr(type_nr)
        if not norm:
            continue
        rec = {
            "artikel": str(artikel),
            "fabrikant": str(fab) if fab else "",
            "type_nr": str(type_nr),
            "omschrijving": str(omsch1) if omsch1 else "",
        }
        by_type_only[norm].append(rec)
        if fab and str(fab) != "XXXXXX":
            by_fab_type[(str(fab), norm)].append(rec)
        n += 1
    wb.close()
    print(f"[load] {n} ProCos articles loaded")
    print(f"[load] unique (fab,type) keys: {len(by_fab_type)}")
    print(f"[load] unique type-only keys:  {len(by_type_only)}")
    return by_fab_type, by_type_only, n


def extract_fcc13_rows():
    cfg = yaml.safe_load(open(ROOT / "extractor" / "config.yaml", encoding="utf-8"))
    print(f"[extract] running pipeline on {PDF_PATH} pages {PDF_PAGES}...")
    result = pipeline_run(PDF_PATH, cfg, PDF_PAGES)
    print(f"[extract] {result.row_count} rows extracted")
    return result.rows


def match_row(row, fab_mapping, by_fab_type, by_type_only):
    """Return (status, candidates, mapped_fab_code, normalized_type)."""
    model = row.model_number or row.order_number or ""
    norm_type = _normalize_typenr(model)
    if not norm_type:
        return "GEEN TYPE NR", [], None, ""

    raw_fab = (row.manufacturer or "").strip().upper()
    mapped_fab = fab_mapping.get(raw_fab)

    # Strict match: (fab, type)
    if mapped_fab:
        hits = by_fab_type.get((mapped_fab, norm_type), [])
        if len(hits) == 1:
            return "MATCH", hits, mapped_fab, norm_type
        if len(hits) > 1:
            return "NIET UNIEK", hits, mapped_fab, norm_type
        # 0 hits with mapping: fallback to type-only
        hits = by_type_only.get(norm_type, [])
        if len(hits) == 1:
            return "MATCH (fab onbekend in mapping, op type alleen)", hits, mapped_fab, norm_type
        if len(hits) > 1:
            return "NIET UNIEK (type alleen)", hits, mapped_fab, norm_type
        return "NIET GEVONDEN", [], mapped_fab, norm_type

    # Fab not in mapping: type-only
    hits = by_type_only.get(norm_type, [])
    if len(hits) == 1:
        return "MATCH (fab niet gemapt, op type alleen)", hits, None, norm_type
    if len(hits) > 1:
        return "NIET UNIEK (type alleen, fab niet gemapt)", hits, None, norm_type
    return "NIET GEVONDEN (fab niet gemapt)", [], None, norm_type


def main():
    by_fab_type, by_type_only, n = load_procos(PROCOS_EXPORT)
    rows = extract_fcc13_rows()

    # Run matches
    results = []
    statuses = defaultdict(int)
    for r in rows:
        status, hits, mapped_fab, norm_type = match_row(
            r, FAB_MAPPING, by_fab_type, by_type_only
        )
        statuses[status] += 1
        proc_art = hits[0]["artikel"] if hits else ""
        proc_omsch = hits[0]["omschrijving"] if hits else ""
        proc_fab_code = hits[0]["fabrikant"] if hits else ""
        results.append({
            "page": r.source_page,
            "device_tag": r.device_tag,
            "qty": r.quantity,
            "description": r.description,
            "manufacturer_pdf": r.manufacturer,
            "model_number_pdf": r.model_number,
            "order_number_pdf": r.order_number,
            "mapped_fabcode": mapped_fab or "",
            "normalized_type": norm_type,
            "status": status,
            "procos_artikel": proc_art,
            "procos_fab": proc_fab_code,
            "procos_omschrijving": proc_omsch,
            "n_hits": len(hits),
        })

    # Print summary
    print()
    print("=== STATUS-OVERZICHT ===")
    for status, count in sorted(statuses.items(), key=lambda x: -x[1]):
        pct = 100 * count / len(rows)
        print(f"  {status:50s}: {count:>3} ({pct:5.1f}%)")
    print(f"  {'TOTAAL':50s}: {len(rows):>3}")

    # Write Excel
    print(f"\n[write] {OUT_REPORT}")
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "FCC-13 match"
    headers = list(results[0].keys()) if results else []
    ws.append(headers)
    for h_cell in ws[1]:
        h_cell.font = openpyxl.styles.Font(bold=True)
    for r in results:
        ws.append([r[h] for h in headers])

    # Summary sheet
    sum_ws = out_wb.create_sheet("Samenvatting")
    sum_ws.append(["Status", "Aantal", "Percentage"])
    for h_cell in sum_ws[1]:
        h_cell.font = openpyxl.styles.Font(bold=True)
    for status, count in sorted(statuses.items(), key=lambda x: -x[1]):
        sum_ws.append([status, count, f"{100*count/len(rows):.1f}%"])
    sum_ws.append(["TOTAAL", len(rows), "100.0%"])

    out_wb.save(OUT_REPORT)
    print(f"[done] {len(results)} rijen geschreven naar {OUT_REPORT}")


if __name__ == "__main__":
    main()
