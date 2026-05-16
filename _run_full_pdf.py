"""End-to-end run voor één PDF:
   1. Classifier → vind stuklijst-pagina's
   2. Extractor → 66/76/... rijen
   3. Schrijf Rauwe Excel, ProCos XLTM, ProCos XML
   4. Match tegen ProCos 86k-export → match-rapport

Output: 4 bestanden in dezelfde map als de PDF, met stem-naam prefix.
"""

from __future__ import annotations

import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "extractor"))

from src.pipeline import run as pipeline_run
from src.writers.procos_writer import write_procos
from src.writers.procos_xml_writer import write_procos_xml
from src.writers.xlsx_writer import write_xlsx

# Classifier import: two-phase swap to avoid src/ name collision
def _classify(pdf_path: str):
    _CLASSIFIER = ROOT / "classifier"
    if str(_CLASSIFIER) not in sys.path:
        sys.path.insert(0, str(_CLASSIFIER))
    _saved = {k: v for k, v in sys.modules.items() if k == "src" or k.startswith("src.")}
    for k in _saved:
        del sys.modules[k]
    try:
        from src.classifier import classify
        runs, _ = classify(pdf_path)
        return [r.pages for r in runs]
    finally:
        for k in list(sys.modules.keys()):
            if k == "src" or k.startswith("src."):
                del sys.modules[k]
        sys.modules.update(_saved)
        try:
            sys.path.remove(str(_CLASSIFIER))
        except ValueError:
            pass


PROCOS_EXPORT = r"C:\Users\JorisMerkx\Downloads\ProCos-export Artikeldata-excl prijzen.xlsx"

FAB_MAPPING = {
    "SIEMENS": "SIEHAA",
    "SIEMENS AG": "SIEHAA",
    "ROCKWELL": "ROCUIT",
    "ROCKWELL AUTOMATION": "ROCUIT",
    "ABB": "ABBROT1",
    "EATON": "EATVEJ",
    "RITTAL": "RITZEV",
    "PHOENIX": "PHOZEV",
    "PHOENIX CONTACT": "PHOZEV",
    "LAPP": "LAPVEL",
    "WEIDMULLER": "WEIHIL",
    "WEIDMUELLER": "WEIHIL",
    "WEIDMÜLLER": "WEIHIL",
    "SCHNEIDER": "SCHHAA",
    "SCHNEIDER ELECTRIC": "SCHHAA",
    "HARTING": "HARBOS",
    "FESTO": "FESDEL",
    "EAO": "EAODOR",
    "HUMMEL": "HUMZEV",
    "FINDER": "FINAMS",
    "SMC": "SMCAMS",
    "SMC PNEUMATICS": "SMCAMS",
    "PMA": "PMAUST",
    "PMA AG": "PMAUST",
    "FIBOX": "FIBESP",
}


def _norm_type(s):
    if not s:
        return ""
    return re.sub(r"[\s\-./()_,]+", "", str(s)).upper()


def load_procos():
    print(f"[load] reading {PROCOS_EXPORT}")
    t0 = time.time()
    wb = openpyxl.load_workbook(PROCOS_EXPORT, read_only=True, data_only=True)
    ws = wb["export"]
    by_fab_type = defaultdict(list)
    by_type_only = defaultdict(list)
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        artikel, _, _, _, fab, type_nr, omsch1, *_ = row
        if not artikel or not type_nr:
            continue
        norm = _norm_type(type_nr)
        if not norm:
            continue
        rec = {
            "artikel": str(artikel),
            "fabrikant": str(fab) if fab else "",
            "type_nr": str(type_nr),
            "omschrijving": str(omsch1) if omsch1 else "",
        }
        by_type_only[norm].append(rec)
        if fab and str(fab) != "XXXXXX":
            by_fab_type[(str(fab), norm)].append(rec)
        n += 1
    wb.close()
    print(f"[load] {n} ProCos articles in {time.time()-t0:.1f}s")
    return by_fab_type, by_type_only


def _candidate_typenrs(row) -> list[str]:
    """Return a list of candidate type/order numbers for matching, in
    priority order. Includes model_number, order_number, and split parts
    of gecombineerde 'type number order number' kolommen uit extra_fields.
    """
    cands: list[str] = []
    if row.model_number:
        cands.append(str(row.model_number))
    if row.order_number and row.order_number != row.model_number:
        cands.append(str(row.order_number))
    # G88000-style combined header in extra_fields
    if isinstance(row.extra_fields, dict):
        for key, val in row.extra_fields.items():
            key_l = str(key).lower()
            if any(t in key_l for t in ("type", "order", "bestelnr", "bstnr", "model")):
                if val:
                    # Combined cell often holds two values separated by newline
                    for part in str(val).split("\n"):
                        part = part.strip()
                        if part and part not in cands:
                            cands.append(part)
    # Deduplicate while preserving order
    seen = set()
    unique = []
    for c in cands:
        n = _norm_type(c)
        if n and n not in seen:
            seen.add(n)
            unique.append(c)
    return unique


def match_one(row, by_fab_type, by_type_only):
    candidates = _candidate_typenrs(row)
    if not candidates:
        return "GEEN TYPE NR", []
    fab_raw = (row.manufacturer or "").strip().upper()
    fab_code = FAB_MAPPING.get(fab_raw)

    # Try each candidate type-nr in priority order
    for cand in candidates:
        norm = _norm_type(cand)
        if not norm:
            continue
        if fab_code:
            hits = by_fab_type.get((fab_code, norm), [])
            if len(hits) == 1:
                return "MATCH", hits
            if len(hits) > 1:
                return "NIET UNIEK", hits
        else:
            hits = by_type_only.get(norm, [])
            if len(hits) == 1:
                return "MATCH (fab niet gemapt, op type alleen)", hits
            if len(hits) > 1:
                return "NIET UNIEK (type alleen, fab niet gemapt)", hits

    return "NIET GEVONDEN" if fab_code else "NIET GEVONDEN (fab niet gemapt)", []


def run(pdf_path: str):
    pdf = Path(pdf_path).resolve()
    out_dir = pdf.parent
    stem = pdf.stem

    cfg = yaml.safe_load(open(ROOT / "extractor" / "config.yaml", encoding="utf-8"))

    print(f"\n=== STAP 1: classifier ===")
    t0 = time.time()
    page_runs = _classify(str(pdf))
    print(f"  detected {len(page_runs)} run(s): "
          f"{[(r[0], r[-1]) for r in page_runs]}  ({time.time()-t0:.1f}s)")

    print(f"\n=== STAP 2: extractor ===")
    t0 = time.time()
    result = pipeline_run(str(pdf), cfg, page_runs)
    print(f"  {result.row_count} rijen ({time.time()-t0:.1f}s)")

    print(f"\n=== STAP 3: schrijf outputs ===")
    raw_xlsx = out_dir / f"{stem}_extractie.xlsx"
    write_xlsx([result], str(raw_xlsx), cfg)
    print(f"  Raw Excel : {raw_xlsx}")

    procos_xltm = out_dir / f"{stem}_procos.xltm"
    write_procos(result, str(procos_xltm))
    print(f"  ProCos XLTM: {procos_xltm}")

    procos_xml = out_dir / f"{stem}_procos.xml"
    write_procos_xml(result, str(procos_xml))
    print(f"  ProCos XML : {procos_xml}")

    print(f"\n=== STAP 4: match tegen ProCos 86k-export ===")
    by_fab_type, by_type_only = load_procos()

    results = []
    statuses = defaultdict(int)
    for r in result.rows:
        status, hits = match_one(r, by_fab_type, by_type_only)
        statuses[status] += 1
        proc = hits[0] if hits else {"artikel": "", "fabrikant": "", "omschrijving": ""}
        results.append({
            "page": r.source_page,
            "device_tag": r.device_tag,
            "qty": r.quantity,
            "description": r.description,
            "manufacturer_pdf": r.manufacturer,
            "model_number_pdf": r.model_number,
            "order_number_pdf": r.order_number,
            "status": status,
            "procos_artikel": proc["artikel"],
            "procos_fabcode": proc["fabrikant"],
            "procos_omschrijving": proc["omschrijving"],
            "n_hits": len(hits),
        })

    print()
    print("=== STATUS-OVERZICHT ===")
    for status, count in sorted(statuses.items(), key=lambda x: -x[1]):
        pct = 100 * count / max(1, len(result.rows))
        print(f"  {status:55s}: {count:>3} ({pct:5.1f}%)")
    print(f"  {'TOTAAL':55s}: {len(result.rows):>3}")

    match_xlsx = out_dir / f"{stem}_match_rapport.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Match"
    headers = list(results[0].keys()) if results else []
    ws.append(headers)
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    for r in results:
        ws.append([r[h] for h in headers])

    sw = wb.create_sheet("Samenvatting")
    sw.append(["Status", "Aantal", "%"])
    for c in sw[1]:
        c.font = openpyxl.styles.Font(bold=True)
    for status, count in sorted(statuses.items(), key=lambda x: -x[1]):
        sw.append([status, count, f"{100*count/max(1,len(result.rows)):.1f}%"])
    sw.append(["TOTAAL", len(result.rows), "100.0%"])

    wb.save(match_xlsx)
    print(f"\n  Match rapport: {match_xlsx}")

    print(f"\n=== KLAAR ===")
    return {
        "pdf": str(pdf),
        "page_runs": page_runs,
        "row_count": result.row_count,
        "statuses": dict(statuses),
        "outputs": {
            "raw_xlsx": str(raw_xlsx),
            "procos_xltm": str(procos_xltm),
            "procos_xml": str(procos_xml),
            "match_xlsx": str(match_xlsx),
        }
    }


if __name__ == "__main__":
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\JorisMerkx\OneDrive - Agyle\Documenten\EKB\Item list poc\data\126-0053 Cabinet Lineator Controller.pdf"
    run(pdf_path)
