"""Excel item-list reader.

Mirrors the PDF extractor's contract: given an xlsx, produce an
ExtractionResult containing CanonicalRow objects ready for the downstream
match step.

Scalability strategy — no klant-specific code paths:
  * Auto-detect the best sheet (skip empty sheets, prefer the one whose
    candidate header has the most synonym hits).
  * Auto-detect the header row (scan the first N rows per sheet, score
    each row by synonym hits + text/numeric ratio).
  * Reuse the existing ``column_mapper`` so any new column-name
    vocabulary works by adding synonyms in ``config.yaml`` (no code change).
  * Unmapped columns flow into ``CanonicalRow.extra_fields`` — Excels with
    42+ columns are fine; we surface what we can map and keep the rest.
  * Reuse ``post_processor`` for normalization (quantity, unicode, etc.).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, List, Optional, Tuple

from openpyxl import load_workbook

from .column_mapper import _normalize, map_columns
from .interfaces import CanonicalRow, ExtractionResult, RawTable
from .post_processor import post_process
from .row_parser import parse_rows


# How many rows to scan when searching for the header row in a sheet.
_HEADER_SCAN_ROWS = 20
# Header candidates must have at least this many non-empty cells.
_MIN_HEADER_CELLS = 3


# ---------------------------------------------------------------------------
# Sheet + header detection
# ---------------------------------------------------------------------------


def _read_all_rows(ws) -> List[List[Any]]:
    """Read every row from ``ws`` as a list-of-lists. Trailing fully-empty
    rows are stripped so very wide-but-mostly-empty sheets don't confuse
    later size checks.
    """
    out: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        out.append(list(row))
    while out and _row_nonempty_count(out[-1]) == 0:
        out.pop()
    return out


def _row_nonempty_count(row: List[Any]) -> int:
    return sum(
        1 for c in row
        if c is not None and (not isinstance(c, str) or c.strip())
    )


def _is_numeric_like(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    s = str(value).strip()
    if not s:
        return False
    try:
        float(s.replace(",", "."))
        return True
    except ValueError:
        return False


def _flatten_synonyms_norm(config: dict) -> set[str]:
    """Return the set of normalized synonym strings for header scoring."""
    cm = (config or {}).get("column_mapping") or {}
    syns = cm.get("synonyms") or {}
    out: set[str] = set()
    for _, lst in syns.items():
        for syn in lst or []:
            if syn:
                norm = _normalize(syn)
                if norm:
                    out.add(norm)
    return out


def _score_header_candidate(row: List[Any], synonyms_norm: set[str]) -> int:
    """Score a row's plausibility as a header. Higher = more header-like.

    Score = (synonym matches × 10) + text-cell count, halved if the row
    looks numeric-dominated (probably a data row). Rows with fewer than
    ``_MIN_HEADER_CELLS`` non-empty cells score 0.
    """
    nonempty = _row_nonempty_count(row)
    if nonempty < _MIN_HEADER_CELLS:
        return 0

    matches = 0
    text_cells = 0
    numeric_cells = 0
    for c in row:
        if c is None:
            continue
        s_raw = str(c).strip()
        if not s_raw:
            continue
        if _is_numeric_like(c):
            numeric_cells += 1
            continue
        text_cells += 1
        if _normalize(s_raw) in synonyms_norm:
            matches += 1

    score = matches * 10 + text_cells
    if numeric_cells > text_cells:
        score = score // 4
    return score


def _find_best_sheet_and_header(
    wb,
    config: dict,
) -> Optional[Tuple[int, str, int, List[Any], List[List[Any]]]]:
    """Pick the (sheet, header_row) combination with the highest score.

    Returns ``(score, sheet_name, header_row_index_0based, header_row,
    data_rows)`` or ``None`` if no sheet has a plausible header.
    """
    synonyms_norm = _flatten_synonyms_norm(config)
    best: Optional[Tuple[int, str, int, List[Any], List[List[Any]]]] = None

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = _read_all_rows(ws)
        if len(rows) < 2:  # need header + at least one data row
            continue

        scan_limit = min(_HEADER_SCAN_ROWS, len(rows) - 1)
        for hdr_idx in range(scan_limit):
            score = _score_header_candidate(rows[hdr_idx], synonyms_norm)
            if score == 0:
                continue
            data_rows = rows[hdr_idx + 1:]
            # Strip trailing all-empty data rows (e.g. styled blank rows).
            while data_rows and _row_nonempty_count(data_rows[-1]) == 0:
                data_rows.pop()
            if not data_rows:
                continue
            cand = (score, sn, hdr_idx, rows[hdr_idx], data_rows)
            if best is None or score > best[0]:
                best = cand

    return best


def _trim_trailing_empty(cells: List[Any]) -> List[Any]:
    out = list(cells)
    while out and (out[-1] is None or str(out[-1]).strip() == ""):
        out.pop()
    return out


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------


def read_xlsx(xlsx_path: str, config: dict) -> ExtractionResult:
    """Read an xlsx file and return an ``ExtractionResult``.

    Never raises on malformed/empty input — returns an empty result with
    an audit note instead, so the caller can surface a friendly message.
    """
    src_name = Path(xlsx_path).name
    result = ExtractionResult(source_pdf=src_name)

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        choice = _find_best_sheet_and_header(wb, config)
    finally:
        wb.close()

    if choice is None:
        result.audit["note"] = "no_plausible_header_row_found"
        return result

    score, sheet_name, header_idx, header_row, data_rows = choice

    # Cap column count to the actual width of the header — many klant-Excels
    # have stale styled columns to the right of real data.
    capped = _trim_trailing_empty(header_row)
    n_cols = len(capped)
    if n_cols == 0:
        result.audit["note"] = "header_row_empty_after_trim"
        return result

    headers: List[str] = [
        "" if header_row[c] is None else str(header_row[c])
        for c in range(n_cols)
    ]

    # Convert data cells to strings for RawTable, truncated to n_cols.
    cell_rows: List[List[str]] = []
    for r in data_rows:
        rr = [
            "" if (c >= len(r) or r[c] is None) else str(r[c])
            for c in range(n_cols)
        ]
        cell_rows.append(rr)

    raw_table = RawTable(
        page_number=1,            # xlsx has no pages — synthetic
        headers=headers,
        rows=cell_rows,
        parser="openpyxl",
        table_index=0,
        n_cols=n_cols,
        n_rows=len(cell_rows),
        bbox=None,
    )

    mappings = map_columns(headers, config)
    row_dicts = parse_rows(raw_table, mappings, config)

    row_index = 0
    for rd in row_dicts:
        if rd.get("_is_section_header"):
            continue
        canonical = CanonicalRow(
            source_pdf=src_name,
            source_page=1,
            source_section=sheet_name,
            row_index=row_index,
            device_tag=rd.get("device_tag") or None,
            quantity=rd.get("quantity"),
            description=rd.get("description") or None,
            manufacturer=rd.get("manufacturer") or None,
            model_number=rd.get("model_number") or None,
            order_number=rd.get("order_number") or None,
            schematic_position=rd.get("schematic_position") or None,
            extra_fields=rd.get("extra_fields") or {},
            raw=list(rd.get("raw") or []),
        )
        canonical = post_process(canonical, config)
        result.rows.append(canonical)
        row_index += 1

    result.audit["xlsx"] = {
        "sheet": sheet_name,
        "header_row_index": header_idx,
        "header_score": score,
        "n_cols": n_cols,
        "n_rows_input": len(data_rows),
        "n_rows_output": len(result.rows),
        "mapped_columns": {
            m.column_index: m.canonical_field
            for m in mappings
            if m.canonical_field
        },
    }
    return result
