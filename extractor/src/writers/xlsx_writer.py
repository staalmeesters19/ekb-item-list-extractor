"""XLSX writer for ExtractionResult."""

from __future__ import annotations

import json
from pathlib import Path
from typing import List

from ..interfaces import CANONICAL_FIELDS, ExtractionResult


def write_xlsx(
    results: List[ExtractionResult],
    output_path: str,
    config: dict,
) -> None:
    """Write one or more ExtractionResult objects to *output_path* as XLSX.

    With ``output.xlsx_sheet_per_pdf: true`` each result gets its own sheet
    (named by the source PDF basename).  All results land on a single sheet
    otherwise.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX output") from exc

    out_cfg = (config or {}).get("output", {}) or {}
    sheet_per_pdf: bool = bool(out_cfg.get("xlsx_sheet_per_pdf", True))
    include_raw: bool = bool(out_cfg.get("include_raw_column", True))
    include_warnings: bool = bool(out_cfg.get("include_warnings_column", True))

    meta_fields = ["source_pdf", "source_page", "source_section", "row_index"]
    extra_field = ["extra_fields"]
    optional = []
    if include_raw:
        optional.append("raw")
    if include_warnings:
        optional.append("warnings")
    fieldnames = meta_fields + CANONICAL_FIELDS + extra_field + optional

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    def _sheet_name(pdf_name: str, idx: int) -> str:
        name = Path(pdf_name).stem[:28]  # Excel sheet names ≤ 31 chars
        return name if name else f"Sheet{idx + 1}"

    def _fill_sheet(ws, rows, fieldnames):
        bold = Font(bold=True)
        ws.append(fieldnames)
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            d = row.to_dict()
            d["extra_fields"] = json.dumps(d.get("extra_fields") or {}, ensure_ascii=False)
            if include_raw:
                d["raw"] = json.dumps(d.get("raw") or [], ensure_ascii=False)
            if include_warnings:
                d["warnings"] = "; ".join(d.get("warnings") or [])
            ws.append([d.get(f) for f in fieldnames])

    if sheet_per_pdf:
        for idx, result in enumerate(results):
            name = _sheet_name(result.source_pdf, idx)
            ws = wb.create_sheet(title=name)
            _fill_sheet(ws, result.rows, fieldnames)
    else:
        ws = wb.create_sheet(title="item_list")
        _fill_sheet(ws, [r for res in results for r in res.rows], fieldnames)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
