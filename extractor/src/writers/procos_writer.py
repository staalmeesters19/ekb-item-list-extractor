"""ProCos-template writer.

Fills the customer-specific ProCos import template (``klantlijst`` sheet)
with the rows from an ``ExtractionResult``.  The output is a macro-enabled
``.xltm`` file: EKB opens it in Excel and clicks the embedded "XML Opslaan"
button to generate the XML that ProCos imports.

The two other sheets in the workbook (``Daten`` and ``XML Ausgabe``)
contain formulas that read from ``klantlijst`` automatically, so we only
ever write to ``klantlijst``.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any, Optional

from ..interfaces import CanonicalRow, ExtractionResult


# Bundled with the package so the writer works regardless of cwd.
TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "ProCosImportStuklijst.xltm"

# klantlijst column layout — header on row 1, data starts row 2.
# (Excel column letter, header label) — for documentation only.
_KLANTLIJST_COLUMNS = (
    "A: Aantal",
    "B: Eenheid",
    "C: Klantartikel",
    "D: Omschrijving",
    "E: Fabrikant",
    "F: Type/bestelnummer",
    "G: toegeleverd",
    "H: ODC code",
    "I: Opmerking",
    "J: EAN code",
)


def _quantity_value(raw: Any) -> Any:
    """Coerce a quantity to int when it represents a whole number, else
    return the original (float / str / None).
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return int(raw)
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        return int(raw) if raw.is_integer() else raw
    return raw


def _type_bestelnummer(row: CanonicalRow) -> str:
    """model_number leidend, fallback op order_number."""
    primary = (row.model_number or "").strip() if row.model_number else ""
    if primary:
        return primary
    return (row.order_number or "").strip() if row.order_number else ""


def _opmerking(row: CanonicalRow) -> str:
    """Section label + warnings, comma-separated. Empty when neither present."""
    parts: list[str] = []
    if row.source_section:
        parts.append(f"[{row.source_section}]")
    if row.warnings:
        parts.append("; ".join(row.warnings))
    return " ".join(parts)


def _build_row_values(row: CanonicalRow) -> list[Any]:
    """Map a CanonicalRow to the 10 klantlijst columns (A..J)."""
    return [
        _quantity_value(row.quantity),                  # A: Aantal
        "Stuks",                                        # B: Eenheid (default)
        (row.device_tag or "").strip() or None,         # C: Klantartikel
        (row.description or "").strip() or None,        # D: Omschrijving
        (row.manufacturer or "").strip() or None,       # E: Fabrikant
        _type_bestelnummer(row) or None,                # F: Type/bestelnummer
        None,                                           # G: toegeleverd (leeg)
        None,                                           # H: ODC code (leeg)
        _opmerking(row) or None,                        # I: Opmerking
        None,                                           # J: EAN code (leeg)
    ]


def write_procos(
    result: ExtractionResult,
    output_path: str,
    config: Optional[dict] = None,
    template_path: Optional[str] = None,
) -> None:
    """Write *result* to a ProCos import template at *output_path*.

    The template is loaded with ``keep_vba=True`` so the embedded macro
    (the "XML Opslaan" button) survives the round-trip.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required for ProCos export") from exc

    src = Path(template_path) if template_path else TEMPLATE_PATH
    if not src.exists():
        raise FileNotFoundError(f"ProCos template not found at {src}")

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    # Copy first, then load the copy. openpyxl read+write on the same path is
    # fine but copying preserves any non-OOXML chunks we don't yet know about.
    shutil.copy2(src, out)

    wb = openpyxl.load_workbook(out, keep_vba=True, data_only=False)
    if "klantlijst" not in wb.sheetnames:
        raise ValueError(
            f"Template at {src} has no 'klantlijst' sheet "
            f"(found: {wb.sheetnames})"
        )
    ws = wb["klantlijst"]

    # Write each CanonicalRow; data starts at row 2 (row 1 is the header).
    for i, row in enumerate(result.rows, start=2):
        values = _build_row_values(row)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=i, column=col_idx, value=value)

    wb.save(out)
    wb.close()
