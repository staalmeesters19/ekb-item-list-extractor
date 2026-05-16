"""Writers for ProCos match-results.

This module produces two Excel reports based on a list of CanonicalRow
objects paired 1:1 with MatchResult objects (from ``..matcher``):

* :func:`write_match_report` -- full match-rapport with a colour-coded
  "Match" sheet and a "Samenvatting" sheet.
* :func:`write_niet_gevonden` -- separate Excel listing only the rows
  that did not produce a clean match, including a human-readable reason.

Both writers depend only on attribute access on the supplied MatchResult
objects (status, procos_artikel, procos_fabcode, procos_omschrijving,
mapped_fab, matched_typenr, n_hits), so the matcher module itself does
not need to be importable here.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, List, Optional, Sequence, Union

from ..interfaces import CanonicalRow


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_FILL_GREEN = "D4F5D4"   # MATCH (any variant)
_FILL_AMBER = "FFE5B4"   # NIET UNIEK (any variant)
_FILL_RED = "FFD4D4"     # NIET GEVONDEN (any variant)
_FILL_GRAY = "E5E5E5"    # GEEN TYPE NR

_MATCH_COLUMNS = [
    "source_page",
    "source_section",
    "device_tag",
    "quantity",
    "description",
    "manufacturer",
    "model_number",
    "order_number",
    "status",
    "procos_artikel",
    "procos_fabcode",
    "procos_omschrijving",
    "matched_typenr",
    "n_hits",
]

_NIET_GEVONDEN_COLUMNS = [
    "source_page",
    "source_section",
    "device_tag",
    "quantity",
    "description",
    "manufacturer",
    "model_number",
    "order_number",
    "status",
    "reden",
]

_REDEN_MAP = {
    "NIET GEVONDEN": "Type-nr niet gevonden in ProCos (fabrikant wel bekend)",
    "NIET GEVONDEN (fab niet gemapt)": "Fabrikant niet in mapping-tabel",
    "NIET UNIEK": "Meerdere matches in ProCos — handmatige keuze nodig",
    "NIET UNIEK (op type alleen)": "Meerdere matches op type-nr alleen",
    "GEEN TYPE NR": "Geen type/bestelnummer beschikbaar",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fill_for_status(status: str) -> Optional[str]:
    """Return the hex fill colour (no leading '#') for a given status, or None."""
    if not status:
        return None
    if status == "GEEN TYPE NR":
        return _FILL_GRAY
    if status.startswith("MATCH"):
        return _FILL_GREEN
    if status.startswith("NIET UNIEK"):
        return _FILL_AMBER
    if status.startswith("NIET GEVONDEN"):
        return _FILL_RED
    return None


def _is_unmatched(status: str) -> bool:
    """Return True if a row should appear in the 'niet gevonden' export."""
    if not status:
        return False
    if status == "GEEN TYPE NR":
        return True
    if status.startswith("NIET GEVONDEN"):
        return True
    if status.startswith("NIET UNIEK"):
        return True
    return False


def _row_values(row: CanonicalRow, match: Any) -> List[Any]:
    """Build the Match-sheet row values for one (CanonicalRow, MatchResult)."""
    return [
        row.source_page,
        row.source_section,
        row.device_tag,
        row.quantity,
        row.description,
        row.manufacturer,
        row.model_number,
        row.order_number,
        getattr(match, "status", "") or "",
        getattr(match, "procos_artikel", "") or "",
        getattr(match, "procos_fabcode", "") or "",
        getattr(match, "procos_omschrijving", "") or "",
        getattr(match, "matched_typenr", "") or "",
        getattr(match, "n_hits", 0),
    ]


def _niet_gevonden_values(row: CanonicalRow, match: Any) -> List[Any]:
    status = getattr(match, "status", "") or ""
    reden = _REDEN_MAP.get(status, "")
    return [
        row.source_page,
        row.source_section,
        row.device_tag,
        row.quantity,
        row.description,
        row.manufacturer,
        row.model_number,
        row.order_number,
        status,
        reden,
    ]


def _ensure_parent(path: Union[str, Path]) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


# ---------------------------------------------------------------------------
# Public writers
# ---------------------------------------------------------------------------

def write_match_report(
    rows: Sequence[CanonicalRow],
    match_results: Sequence[Any],
    output_path: Union[str, Path],
    config: Optional[dict] = None,
) -> None:
    """Write the full match-rapport (Match + Samenvatting) to *output_path*.

    Parameters
    ----------
    rows
        CanonicalRow objects (one per extracted item-list row).
    match_results
        MatchResult objects of equal length, paired 1:1 with *rows*.
    output_path
        Target xlsx path (str or pathlib.Path).  Parent directories are
        created if missing.
    config
        Currently unused; reserved for future options.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError("openpyxl is required for match-report output") from exc

    if len(rows) != len(match_results):
        raise ValueError(
            f"rows and match_results must have equal length "
            f"({len(rows)} vs {len(match_results)})"
        )

    _ = config  # reserved

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)

    # -- Sheet 1: Match -----------------------------------------------------
    ws = wb.create_sheet(title="Match")
    ws.append(_MATCH_COLUMNS)
    for cell in ws[1]:
        cell.font = bold

    for row, match in zip(rows, match_results):
        values = _row_values(row, match)
        ws.append(values)
        fill_hex = _fill_for_status(getattr(match, "status", "") or "")
        if fill_hex:
            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            last_row = ws.max_row
            for col_idx in range(1, len(_MATCH_COLUMNS) + 1):
                ws.cell(row=last_row, column=col_idx).fill = fill

    # -- Sheet 2: Samenvatting ---------------------------------------------
    ws2 = wb.create_sheet(title="Samenvatting")

    total = len(rows)
    source_pdf = rows[0].source_pdf if rows else ""

    # Count statuses
    status_counts: dict[str, int] = {}
    match_count = 0
    for m in match_results:
        st = getattr(m, "status", "") or ""
        status_counts[st] = status_counts.get(st, 0) + 1
        if st.startswith("MATCH"):
            match_count += 1

    match_pct = (match_count / total * 100.0) if total else 0.0

    # Top section (3 rows)
    ws2.append([f"PDF: {source_pdf}"])
    ws2.append([f"Rijen: {total}"])
    ws2.append([f"Match-percentage: {match_pct:.1f}%"])

    # Blank spacer
    ws2.append([])

    # Status table header
    header_row_idx = ws2.max_row + 1
    ws2.append(["Status", "Aantal", "%"])
    for cell in ws2[header_row_idx]:
        cell.font = bold

    # Sorted by count desc, then status asc for stability
    sorted_items = sorted(
        status_counts.items(), key=lambda kv: (-kv[1], kv[0])
    )
    for status, count in sorted_items:
        pct = (count / total * 100.0) if total else 0.0
        ws2.append([status, count, f"{pct:.1f}%"])

    # Totaal
    ws2.append(["TOTAAL", total, "100.0%"])

    _ensure_parent(output_path)
    wb.save(str(output_path))


def write_niet_gevonden(
    rows: Sequence[CanonicalRow],
    match_results: Sequence[Any],
    output_path: Union[str, Path],
    config: Optional[dict] = None,
) -> None:
    """Write a separate Excel with only the unmatched rows.

    Includes rows whose status starts with ``"NIET GEVONDEN"`` or
    ``"NIET UNIEK"``, or equals ``"GEEN TYPE NR"``.  When every row is a
    clean match, a placeholder row is emitted so the file is still valid.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError("openpyxl is required for match-report output") from exc

    if len(rows) != len(match_results):
        raise ValueError(
            f"rows and match_results must have equal length "
            f"({len(rows)} vs {len(match_results)})"
        )

    _ = config  # reserved

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Niet gevonden")

    bold = Font(bold=True)
    ws.append(_NIET_GEVONDEN_COLUMNS)
    for cell in ws[1]:
        cell.font = bold

    any_written = False
    for row, match in zip(rows, match_results):
        status = getattr(match, "status", "") or ""
        if not _is_unmatched(status):
            continue
        ws.append(_niet_gevonden_values(row, match))
        any_written = True

    if not any_written:
        placeholder = [""] * len(_NIET_GEVONDEN_COLUMNS)
        placeholder[0] = "Geen rijen — alles is gematched."
        ws.append(placeholder)

    _ensure_parent(output_path)
    wb.save(str(output_path))
