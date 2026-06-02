"""Match extracted PDF rows against a ProCos artikel-database.

This module exposes a small public API for loading the ProCos export
(an ~86k-row Excel file) into a lookup structure and matching each
CanonicalRow produced by the extractor pipeline against it.

Public surface:
    - MatchResult       : dataclass with one row's match outcome
    - load_procos_db()  : load Excel export into lookup dict
    - match_rows()      : run match for a list of CanonicalRow
    - summarize()       : aggregate statistics over MatchResults

The fab-mapping (PDF-fabrikant-text -> ProCos fabcode) is NOT defined
here -- it is passed in by the caller, so this module stays pure.
"""

from __future__ import annotations

import re
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

from .interfaces import CanonicalRow


# ----------------------------------------------------------------------
# Public dataclasses
# ----------------------------------------------------------------------

@dataclass
class MatchResult:
    """Outcome for one CanonicalRow matched against the ProCos database."""
    status: str                # see docstring of match_rows for full list
    procos_artikel: str = ""   # "" if no hit
    procos_fabcode: str = ""   # ProCos fabrikantcode that was hit ("" if none)
    procos_omschrijving: str = ""  # ProCos's own description ("" if no hit)
    mapped_fab: str = ""       # the mapped fab code used for lookup
    matched_typenr: str = ""   # candidate type-nr that matched / was tried last
    n_hits: int = 0            # number of candidate hits (0/1/many)
    matched_route: str = ""    # which cascade route hit: "v2:fab+type", "v2:fab+artcode",
                               #   "v2:fab+bestelnr", "v2:type-only", "v1:fab+type",
                               #   "v1:type-only" — empty when no match.


# ----------------------------------------------------------------------
# Private helpers
# ----------------------------------------------------------------------

_NORM_RE = re.compile(r"[\s\-./()_,\\]+")  # includes backslash per Gino's spec


def _norm_type(s: str | None) -> str:
    """Strip whitespace / dashes / dots / slashes / backslashes / parens /
    underscores / commas and uppercase. Mirrors ProCos's match-time
    normalization (per Import referenties HEADER: "Uitgesloten karakters")."""
    if not s:
        return ""
    return _NORM_RE.sub("", str(s)).upper()


# Common corporate suffixes to strip when matching fab names — these only
# appear in klant data (e.g. "SIEMENS AG") and never in v1/v2 fab tables.
_FAB_SUFFIX_STRIP_RE = re.compile(
    r"\s+(AG|GMBH|B\.?\s?V\.?|BV|N\.?\s?V\.?|NV|INC\.?|LTD\.?|LLC|"
    r"ELECTRIC|ELECTRICAL|ELECTRONICS|CONTACT|AUTOMATION|GROUP|HOLDING|"
    r"CO\.?|COMPANY|CORP\.?|CORPORATION|SE)$",
    re.IGNORECASE,
)


def _fab_variants(s: str | None) -> list[str]:
    """Return progressive normalisations of a klant fab-name to try.

    Mirrors how the same brand appears under different klant conventions:
      "SIEMENS AG"          -> ["SIEMENS AG", "SIEMENS"]
      "Phoenix Contact"     -> ["PHOENIX CONTACT", "PHOENIX"]
      "Schneider Electric"  -> ["SCHNEIDER ELECTRIC", "SCHNEIDER"]
    First entry is always the uppercased input; suffix-stripped variants
    follow. Caller probes lookups in order, taking the first hit.
    """
    if not s:
        return []
    base = str(s).strip().upper()
    if not base:
        return []
    variants = [base]
    # Iteratively strip recognized suffixes (some names have two).
    cur = base
    for _ in range(3):
        stripped = _FAB_SUFFIX_STRIP_RE.sub("", cur).strip()
        if stripped == cur or not stripped:
            break
        cur = stripped
        if cur not in variants:
            variants.append(cur)
    return variants


def _candidate_typenrs(row: CanonicalRow) -> list[str]:
    """Priority list of type-nr candidates to try for the given row.

    Order:
        1. row.model_number
        2. row.order_number (if different from model_number)
        3. Newline-split parts of any extra_fields value whose key looks
           like a type / order / model / bestelnr / bstnr column.

    Duplicates (by normalized form) are removed, preserving first occurrence.
    """
    cands: list[str] = []
    if row.model_number:
        cands.append(str(row.model_number))
    if row.order_number and row.order_number != row.model_number:
        cands.append(str(row.order_number))

    if isinstance(row.extra_fields, dict):
        for key, val in row.extra_fields.items():
            key_l = str(key).lower()
            if any(t in key_l for t in ("type", "order", "bestelnr", "bstnr", "model")):
                if val:
                    for part in str(val).split("\n"):
                        part = part.strip()
                        if part:
                            cands.append(part)

    seen: set[str] = set()
    unique: list[str] = []
    for c in cands:
        n = _norm_type(c)
        if n and n not in seen:
            seen.add(n)
            unique.append(c)
    return unique


def _open_workbook(source: Any):
    """Open the ProCos xlsx from either a path-like or a bytes-like source.

    openpyxl can read from a file path or a file-like object. For raw bytes
    (e.g. from Streamlit's uploader) we wrap them in BytesIO. Anything else
    is written to a NamedTemporaryFile first.
    """
    from io import BytesIO

    if isinstance(source, (str, Path)):
        return openpyxl.load_workbook(str(source), read_only=True, data_only=True)

    if isinstance(source, (bytes, bytearray)):
        return openpyxl.load_workbook(BytesIO(source), read_only=True, data_only=True)

    # File-like object (has .read) -- e.g. Streamlit UploadedFile
    if hasattr(source, "read"):
        data = source.read()
        if isinstance(data, str):
            data = data.encode("utf-8", errors="ignore")
        # Reset the cursor if possible, in case the caller wants to re-read
        if hasattr(source, "seek"):
            try:
                source.seek(0)
            except Exception:
                pass
        return openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)

    # Last resort: dump to a temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(bytes(source))
        tmp_path = Path(tmp.name)
    return openpyxl.load_workbook(str(tmp_path), read_only=True, data_only=True)


# ----------------------------------------------------------------------
# Public functions
# ----------------------------------------------------------------------

def load_procos_db(source: Any) -> dict:
    """Load the ProCos export into a lookup dict.

    Args:
        source: filesystem path (str/Path) or bytes-like / file-like object.
                Must point to an xlsx with a sheet named "export" and at
                least these columns in this order:
                  Artikel | BG | Voorraad | Artikelsoort | Fabrikant |
                  Type nr. | Omschrijving 1 | ... (15 cols total)

    Returns:
        A dict with keys:
            "by_fab_type":  {(fab_code, normalized_type): [record_dicts]}
            "by_type_only": {normalized_type: [record_dicts]}
            "n_rows":       int (number of valid articles loaded)

        Each record_dict is:
            {"artikel", "fabrikant", "type_nr", "omschrijving"}

    Articles with fabrikant == "XXXXXX" are excluded from by_fab_type
    (XXXXXX is ProCos's placeholder for "unknown"), but still appear in
    by_type_only.
    """
    wb = _open_workbook(source)
    try:
        ws = wb["export"]

        by_fab_type: dict[tuple[str, str], list[dict]] = defaultdict(list)
        by_type_only: dict[str, list[dict]] = defaultdict(list)
        n = 0

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                # header row
                continue
            if not row or len(row) < 7:
                continue

            artikel = row[0]
            fab = row[4]
            type_nr = row[5]
            omsch1 = row[6]

            if not artikel or not type_nr:
                continue
            norm = _norm_type(type_nr)
            if not norm:
                continue

            rec = {
                "artikel": str(artikel),
                "fabrikant": str(fab) if fab else "",
                "type_nr": str(type_nr),
                "omschrijving": str(omsch1) if omsch1 else "",
            }
            by_type_only[norm].append(rec)
            if fab and str(fab) != "XXXXXX":
                by_fab_type[(str(fab), norm)].append(rec)
            n += 1
    finally:
        wb.close()

    # Build aux index for wildcard fab-lookup (Fase C: Adressen "SIE%").
    by_type_to_fabs: dict[str, dict[str, list[dict]]] = defaultdict(dict)
    for (fab_code, type_norm), recs in by_fab_type.items():
        by_type_to_fabs[type_norm][fab_code] = recs

    return {
        "by_fab_type":     dict(by_fab_type),
        "by_type_to_fabs": dict(by_type_to_fabs),
        "by_type_only":    dict(by_type_only),
        "n_rows":          n,
    }


def match_rows(
    rows: Iterable[CanonicalRow],
    db: dict,
    fab_mapping: dict[str, str],
) -> list[MatchResult]:
    """Match each CanonicalRow against the ProCos database.

    Status values produced:
        "MATCH"                       -- unique hit on (fab, type)
        "NIET UNIEK"                  -- >1 hit on (fab, type)
        "NIET GEVONDEN"               -- no hit but fab was mapped
        "MATCH (op type alleen)"      -- unique hit on type-only (fab not mapped)
        "NIET UNIEK (op type alleen)" -- >1 hit on type-only (fab not mapped)
        "NIET GEVONDEN (fab niet gemapt)" -- no hit, fab not mapped
        "GEEN TYPE NR"                -- row had no usable type-nr candidates

    Args:
        rows: list of CanonicalRow objects (from extractor.pipeline)
        db: result of load_procos_db()
        fab_mapping: dict mapping uppercase PDF-fabrikant-tekst to ProCos
                     fabcode, e.g. {"SIEMENS": "SIEHAA", "EATON": "EATVEJ"}.

    Returns:
        List of MatchResult, in the same order as the input rows.
    """
    by_fab_type = db["by_fab_type"]
    by_type_only = db["by_type_only"]

    results: list[MatchResult] = []
    for row in rows:
        results.append(_match_one(row, by_fab_type, by_type_only, fab_mapping))
    return results


def _match_one(
    row: CanonicalRow,
    by_fab_type: dict[tuple[str, str], list[dict]],
    by_type_only: dict[str, list[dict]],
    fab_mapping: dict[str, str],
) -> MatchResult:
    candidates = _candidate_typenrs(row)
    if not candidates:
        return MatchResult(status="GEEN TYPE NR")

    fab_raw = (row.manufacturer or "").strip().upper()
    fab_code = fab_mapping.get(fab_raw, "")

    last_norm = ""
    for cand in candidates:
        norm = _norm_type(cand)
        if not norm:
            continue
        last_norm = norm

        if fab_code:
            hits = by_fab_type.get((fab_code, norm), [])
            if len(hits) == 1:
                h = hits[0]
                return MatchResult(
                    status="MATCH",
                    procos_artikel=h["artikel"],
                    procos_fabcode=h["fabrikant"],
                    procos_omschrijving=h["omschrijving"],
                    mapped_fab=fab_code,
                    matched_typenr=cand,
                    n_hits=1,
                )
            if len(hits) > 1:
                h = hits[0]
                return MatchResult(
                    status="NIET UNIEK",
                    procos_artikel=h["artikel"],
                    procos_fabcode=h["fabrikant"],
                    procos_omschrijving=h["omschrijving"],
                    mapped_fab=fab_code,
                    matched_typenr=cand,
                    n_hits=len(hits),
                )
        else:
            hits = by_type_only.get(norm, [])
            if len(hits) == 1:
                h = hits[0]
                return MatchResult(
                    status="MATCH (op type alleen)",
                    procos_artikel=h["artikel"],
                    procos_fabcode=h["fabrikant"],
                    procos_omschrijving=h["omschrijving"],
                    mapped_fab="",
                    matched_typenr=cand,
                    n_hits=1,
                )
            if len(hits) > 1:
                h = hits[0]
                return MatchResult(
                    status="NIET UNIEK (op type alleen)",
                    procos_artikel=h["artikel"],
                    procos_fabcode=h["fabrikant"],
                    procos_omschrijving=h["omschrijving"],
                    mapped_fab="",
                    matched_typenr=cand,
                    n_hits=len(hits),
                )

    return MatchResult(
        status="NIET GEVONDEN" if fab_code else "NIET GEVONDEN (fab niet gemapt)",
        mapped_fab=fab_code,
        matched_typenr=last_norm,
        n_hits=0,
    )


# ----------------------------------------------------------------------
# Klant-code detectie uit filename + sheet-naam
# ----------------------------------------------------------------------
#
# Heuristiek: zoek herkenbare klant-fragmenten in de filename of sheet-naam.
# Voor elke bekende klant-code een lijst van fragmenten die we kunnen
# herkennen (klant-namen of bedrijfsmerken die in EKB-werkstukken
# vóórkomen). Match insensitive — uppercase haystack vs uppercase fragments.
#
# Niet-klant-specifieke regel: als niets matcht → caller defaults to
# EKBBEVIP (Gino's "Variabelen voor klant" header) of None (skip klant-ref).
_KLANT_FRAGMENTS: dict[str, tuple[str, ...]] = {
    "JBTAMS":    ("JBT", "JBTAMS"),
    "BOSBOX":    ("BOSCH", "BOSBOX", "REXROTH"),
    "EPLAPE":    ("EPLAN", "EPLAPE"),
    "HOUVLA":    ("HOUVLA",),
    "NEDHER1":   ("NEDAP", "NEDHER"),
    "WEIVEN":    ("WEIDMULLER", "WEIVEN"),
    "SELBEV":    ("SELBEV",),
    "STSVEL":    ("STSVEL",),
    "SEEENK":    ("SEEENK",),
    "PANERM":    ("PANASONIC", "PANERM"),
    "SPGBOX":    ("SPGBOX",),
    "EKBDRASPB": ("EKBDRA", "DRACHTEN"),
    "LISOUD":    ("LISOUD",),
    "SMIEIN":    ("SMIEIN",),
    "EKBSOMIP":  ("EKBSOM", "SOMEREN"),
    "EKBBEVIP":  ("EKBBEV", "BEVERWIJK"),
}


def detect_klant_code(*hints: str | None) -> str | None:
    """Try to identify the klant code from filename / sheet-name / etc.

    Returns the matched klant code (e.g. ``"JBTAMS"``), or ``None`` if no
    fragment matched. The caller decides whether ``None`` means "fall back
    to EKBBEVIP default" or "skip the klant-ref cascade step".
    """
    haystack = " ".join(h or "" for h in hints).upper()
    if not haystack.strip():
        return None
    for code, fragments in _KLANT_FRAGMENTS.items():
        if any(frag in haystack for frag in fragments):
            return code
    return None


# ----------------------------------------------------------------------
# v2 — Cascade against the new 232k Artikellijst export
# ----------------------------------------------------------------------
#
# The new ProCos export has a different shape than the legacy 86k file:
#
#   col 1: Artikel EKB       -- the match target ("ABB.1MRK000863-AS")
#   col 2: Omschrijving intern
#   col 3: Type              -- typenr text variant
#   col 4: Eannr             -- EAN/barcode (90.7% filled)
#   col 5: Art code          -- leveranciers-artikelcode
#   col 6: Bestelnr lev      -- leveranciers-bestelnr (98.6%)
#   col 7: Fabrikaat         -- LEESBARE naam ("ABB", "Siemens") — no fab_mapping needed
#   col 8-11: Leverancierscode + naam + sub
#
# Build three lookup tables for the cascade:
#   by_fab_type     : (FABRIKAAT_norm, TYPE_norm)      -> [rec]
#   by_fab_artcode  : (FABRIKAAT_norm, ARTCODE_norm)   -> [rec]
#   by_fab_bestelnr : (FABRIKAAT_norm, BESTELNR_norm)  -> [rec]
#   by_type_only    : TYPE_norm                        -> [rec]   (fallback)
#
# We deliberately do NOT build by_ean yet — needs EAN-extraction from
# PDFs/Excels first (Fase D).


def _norm_fab(s: Any) -> str:
    """Uppercase + strip — Fabrikaat is plain text ('ABB', 'Siemens')."""
    if not s:
        return ""
    return str(s).strip().upper()


def load_procos_db_v2(source: Any) -> dict:
    """Load the *new* 232k ProCos Artikellijst into a v2 lookup dict.

    Args:
        source: filesystem path (str/Path) or bytes-like / file-like object.
                Must be an xlsx with the columns described above, header row 1.

    Returns:
        Dict with keys:
            "by_fab_type":     {(fab_norm, type_norm): [rec]}
            "by_fab_artcode":  {(fab_norm, artcode_norm): [rec]}
            "by_fab_bestelnr": {(fab_norm, bestelnr_norm): [rec]}
            "by_type_only":    {type_norm: [rec]}
            "by_artikel_ekb":  {artikel_ekb: rec}  -- reverse lookup for klant-ref hits
            "n_rows":          int
    """
    wb = _open_workbook(source)
    try:
        # First non-empty sheet
        ws = wb[wb.sheetnames[0]]

        by_fab_type: dict[tuple[str, str], list[dict]] = defaultdict(list)
        by_fab_artcode: dict[tuple[str, str], list[dict]] = defaultdict(list)
        by_fab_bestelnr: dict[tuple[str, str], list[dict]] = defaultdict(list)
        by_type_only: dict[str, list[dict]] = defaultdict(list)
        by_artikel_ekb: dict[str, dict] = {}
        n = 0

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                continue  # header
            if not row or len(row) < 7:
                continue

            artikel = row[0]
            omsch = row[1]
            type_nr = row[2]
            # ean = row[3]  # reserved for Fase D
            art_code = row[4]
            bestelnr = row[5]
            fab = row[6]

            if not artikel:
                continue

            fab_norm = _norm_fab(fab)
            type_norm = _norm_type(type_nr)
            artcode_norm = _norm_type(art_code)
            bestelnr_norm = _norm_type(bestelnr)

            # At least one of the three lookup keys must be present.
            if not (type_norm or artcode_norm or bestelnr_norm):
                continue

            rec = {
                "artikel": str(artikel),
                "fabrikant": str(fab) if fab else "",
                "type_nr": str(type_nr) if type_nr else "",
                "omschrijving": str(omsch) if omsch else "",
            }

            if fab_norm and type_norm:
                by_fab_type[(fab_norm, type_norm)].append(rec)
            if fab_norm and artcode_norm:
                by_fab_artcode[(fab_norm, artcode_norm)].append(rec)
            if fab_norm and bestelnr_norm:
                by_fab_bestelnr[(fab_norm, bestelnr_norm)].append(rec)
            # Fallback: type-only (covers articles without Fabrikaat AND
            # provides a no-fab safety-net for rows where row.manufacturer
            # is missing).
            if type_norm:
                by_type_only[type_norm].append(rec)
            # Reverse: EKB-artikelcode -> full record (used to enrich
            # klant-ref hits with fab/omschrijving/typenr).
            by_artikel_ekb[str(artikel)] = rec

            n += 1
    finally:
        wb.close()

    return {
        "by_fab_type":     dict(by_fab_type),
        "by_fab_artcode":  dict(by_fab_artcode),
        "by_fab_bestelnr": dict(by_fab_bestelnr),
        "by_type_only":    dict(by_type_only),
        "by_artikel_ekb":  by_artikel_ekb,
        "n_rows":          n,
    }


# ----------------------------------------------------------------------
# Import referenties (Gino's HEADER + Eenheden + Adressen sections)
# ----------------------------------------------------------------------
#
# Three sections in one sheet, separated by "*** Sectie ***" markers:
#
#   HEADER:
#     "Variabelen voor klant:"   -> klant-code default
#     "Uitgesloten karakters"    -> regex char-class for typenr normalization
#     "Conversielijst"           -> klant-referentielijst-id
#   *** Eenheden ***
#     klant-eenheid              -> ProCos-eenheid (e.g. "MTR" -> "Meter")
#   *** Adressen ***
#     klant-fabrikant-naam       -> ProCos-fabcode (or wildcard "SIE%")
#
# The Adressen section replaces our hardcoded 24-entry fab_mapping with
# 738 entries — many of which use the "%" wildcard suffix (`SIE%` matches
# any leverancierscode starting with SIE).


def _split_imp_ref_chars(raw: str) -> str:
    """Convert Gino's `Uitgesloten karakters` string into a char-class
    pattern suitable for ``re.compile``. The value is a sequence of
    characters that should be stripped — preserve them verbatim but
    escape regex metacharacters.
    """
    safe: list[str] = []
    for ch in (raw or ""):
        if ch in r"\^]-":
            safe.append("\\" + ch)
        elif ch == " ":
            safe.append(r"\s")  # space -> any whitespace
        else:
            safe.append(ch)
    return "".join(safe)


def load_import_referenties(source: Any) -> dict:
    """Load Gino's Import referenties xlsx into structured config.

    Returns:
        {
            "default_klant_code": "EKBBEVIP",
            "uitgesloten_karakters": "- ,/\\_().",
            "eenheden":  {"st": "Stuks", "MTR": "Meter", ...},
            "adressen":  {
                # klant-fab-name (uppercase, stripped) -> mapping spec
                "SIEMENS": {"raw": "SIE%", "prefix": "SIE", "exact": None},
                "EAO":     {"raw": "EAODOR", "prefix": None, "exact": "EAODOR"},
                ...
            },
            "n_adressen": int,
            "n_eenheden": int,
        }
    """
    wb = _open_workbook(source)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    default_klant = ""
    uitgesloten = ""
    eenheden: dict[str, str] = {}
    adressen: dict[str, dict] = {}

    section = "HEADER"
    for row in rows:
        if not row:
            continue
        # Detect section markers in any cell.
        marker = None
        for v in row:
            if v and isinstance(v, str):
                s = v.strip()
                if s.startswith("***") and s.endswith("***"):
                    marker = s.strip("*").strip()
                    break
        if marker:
            section = marker
            continue

        key = row[0]
        val = row[1] if len(row) > 1 else None
        if key is None and val is None:
            continue
        key_s = str(key or "").strip()
        val_s = str(val or "").strip()

        if section == "HEADER":
            if key_s.lower().startswith("variabelen voor klant"):
                default_klant = val_s
            elif key_s.lower().startswith("uitgesloten karakters"):
                uitgesloten = val_s
            # Conversielijst row is informational; not used for matching.
        elif section.lower().startswith("eenheden"):
            if key_s:
                eenheden[key_s] = val_s
        elif section.lower().startswith("adressen"):
            if not key_s:
                continue
            kc = key_s.upper()
            # Parse mapping value: ends with % -> wildcard prefix; else exact
            if val_s.endswith("%"):
                prefix = val_s[:-1]
                adressen[kc] = {"raw": val_s, "prefix": prefix.upper(), "exact": None}
            else:
                adressen[kc] = {"raw": val_s, "prefix": None, "exact": val_s.upper()}

    return {
        "default_klant_code": default_klant,
        "uitgesloten_karakters": uitgesloten,
        "eenheden": eenheden,
        "adressen": adressen,
        "n_adressen": len(adressen),
        "n_eenheden": len(eenheden),
    }


# ----------------------------------------------------------------------
# Klant referentielijst (Gino's 45k klant-artikel -> EKB-artikel mapping)
# ----------------------------------------------------------------------


def load_klant_referentielijsten(source: Any) -> dict:
    """Load the Klant referentielijsten xlsx into a per-klant lookup.

    Source xlsx columns (header rij 1):
        Klantcode | Klantnaam | Referentielijst | Artikel EKB | Artikel klant

    Returns:
        {
            klant_code_upper: {
                artikel_klant_norm: artikel_ekb,
            },
            ...
        }

    Lookup is normalized via ``_norm_type`` (the same normalization used
    for type-nrs) so klant-artikelcodes with dashes/dots/spaces are
    matchable regardless of formatting.
    """
    wb = _open_workbook(source)
    try:
        ws = wb[wb.sheetnames[0]]
        out: dict[str, dict[str, str]] = defaultdict(dict)
        n = 0
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                continue
            if not row or len(row) < 5:
                continue
            klant_code = row[0]
            artikel_ekb = row[3]
            artikel_klant = row[4]
            if not (klant_code and artikel_ekb and artikel_klant):
                continue
            kc = str(klant_code).strip().upper()
            ak_norm = _norm_type(artikel_klant)
            if not ak_norm:
                continue
            # First occurrence wins (no overwriting).
            if ak_norm not in out[kc]:
                out[kc][ak_norm] = str(artikel_ekb)
                n += 1
    finally:
        wb.close()
    out["_meta"] = {"n_rows": n, "n_klants": len(out)}
    return dict(out)


def _try_klant_ref(
    row: CanonicalRow,
    klant_db: dict | None,
    klant_code: str | None,
    by_artikel_ekb: dict[str, dict] | None,
    fab_norm: str,
) -> MatchResult | None:
    """Try cascade step 0: direct klant-artikel -> EKB-artikel lookup.

    Returns a populated MatchResult when the klant-ref lookup hits, otherwise
    None (so the caller can continue with the fab+type cascade).
    """
    if not klant_db or not klant_code:
        return None
    klant_map = klant_db.get(klant_code.upper())
    if not klant_map:
        return None

    # The klant-artikel code typically lives in row.order_number for our
    # xlsx-reader's mapping (synonym "artikel" -> order_number), but PDFs
    # may have put it in model_number or even device_tag. Try all sensibly
    # promising candidates.
    raw_cands: list[str] = []
    for v in (row.order_number, row.model_number, row.device_tag):
        if v:
            raw_cands.append(str(v))
    # Also fall back to the typenr-candidates helper for completeness.
    for v in _candidate_typenrs(row):
        if v and v not in raw_cands:
            raw_cands.append(v)

    for raw in raw_cands:
        ak_norm = _norm_type(raw)
        if not ak_norm:
            continue
        ekb = klant_map.get(ak_norm)
        if not ekb:
            continue
        # Enrich with the canonical record from the 232k v2 DB if we have it.
        enrich = (by_artikel_ekb or {}).get(ekb, {})
        return MatchResult(
            status="MATCH",
            procos_artikel=ekb,
            procos_fabcode=enrich.get("fabrikant", ""),
            procos_omschrijving=enrich.get("omschrijving", ""),
            mapped_fab=fab_norm,
            matched_typenr=raw,
            n_hits=1,
            matched_route="v2:klant-ref",
        )
    return None


def _match_one_v2(
    row: CanonicalRow,
    db_v2: dict,
    klant_db: dict | None = None,
    klant_code: str | None = None,
) -> MatchResult:
    """Cascade match against the v2 (232k) database, with optional
    klant-referentielijst as cascade step 0.

    Steps (first unique hit wins):
      0. (klant_code, klant_artikel) -> EKB-artikel    [if klant context known]
      3. (fabrikaat, type)
      4. (fabrikaat, art code)
      5. (fabrikaat, bestelnr lev)
      6. type-only (when fab missing or all fab-routes empty)

    A non-unique hit on steps 3-6 is remembered as the best result so far
    but the cascade keeps trying — a later route may produce a unique
    hit that we should prefer.
    """
    fab_norm = _norm_fab(row.manufacturer)

    # Step 0 — klant-ref lookup. Highest-confidence: the klant's own
    # artikelcode maps directly to an EKB-artikel.
    kr_hit = _try_klant_ref(
        row, klant_db, klant_code,
        db_v2.get("by_artikel_ekb"), fab_norm,
    )
    if kr_hit is not None:
        return kr_hit

    candidates = _candidate_typenrs(row)
    if not candidates:
        return MatchResult(status="GEEN TYPE NR")

    by_fab_type     = db_v2.get("by_fab_type") or {}
    by_fab_artcode  = db_v2.get("by_fab_artcode") or {}
    by_fab_bestelnr = db_v2.get("by_fab_bestelnr") or {}
    by_type_only    = db_v2.get("by_type_only") or {}

    best_ambiguous: MatchResult | None = None
    last_norm = ""

    def _ambig(hits, cand, route) -> MatchResult:
        h = hits[0]
        return MatchResult(
            status="NIET UNIEK", procos_artikel=h["artikel"],
            procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
            mapped_fab=fab_norm, matched_typenr=cand, n_hits=len(hits),
            matched_route=route,
        )

    def _unique(hits, cand, route, status="MATCH") -> MatchResult:
        h = hits[0]
        return MatchResult(
            status=status, procos_artikel=h["artikel"],
            procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
            mapped_fab=fab_norm, matched_typenr=cand, n_hits=1,
            matched_route=route,
        )

    for cand in candidates:
        norm = _norm_type(cand)
        if not norm:
            continue
        last_norm = norm

        if fab_norm:
            # Stap 3 → 4 → 5 — try every fab-bound route, keep searching
            # for a unique hit even after a NIET UNIEK on an earlier route.
            for route, table in (
                ("v2:fab+type",     by_fab_type),
                ("v2:fab+artcode",  by_fab_artcode),
                ("v2:fab+bestelnr", by_fab_bestelnr),
            ):
                hits = table.get((fab_norm, norm), [])
                if len(hits) == 1:
                    return _unique(hits, cand, route)
                if len(hits) > 1 and best_ambiguous is None:
                    best_ambiguous = _ambig(hits, cand, route)

        # Stap 6: type-only fallback — only treated as MATCH when unique.
        hits = by_type_only.get(norm, [])
        if len(hits) == 1:
            return _unique(hits, cand, "v2:type-only", status="MATCH (op type alleen)")
        if len(hits) > 1 and best_ambiguous is None:
            h = hits[0]
            best_ambiguous = MatchResult(
                status="NIET UNIEK (op type alleen)", procos_artikel=h["artikel"],
                procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
                mapped_fab=fab_norm, matched_typenr=cand, n_hits=len(hits),
                matched_route="v2:type-only",
            )

    if best_ambiguous is not None:
        return best_ambiguous

    return MatchResult(
        status="NIET GEVONDEN" if fab_norm else "NIET GEVONDEN (fab niet gemapt)",
        mapped_fab=fab_norm,
        matched_typenr=last_norm,
    )


def _resolve_fab_v1_adressen(
    fab_raw: str | None,
    adressen: dict | None,
    fab_mapping_legacy: dict[str, str] | None,
) -> dict | None:
    """Resolve a klant fab-name to a ProCos fab-code or fab-code-prefix.

    Tries (in order, first hit wins):
      1. Adressen exact match on each fab-name variant (suffix-stripped)
      2. Legacy hardcoded fab_mapping (24 entries) on each variant
    Returns ``{"exact": code, "prefix": None}`` or ``{"exact": None, "prefix": pfx}``,
    or None when nothing matched.
    """
    if not fab_raw:
        return None
    variants = _fab_variants(fab_raw)
    if not variants:
        return None
    if adressen:
        for v in variants:
            entry = adressen.get(v)
            if entry:
                return entry
    if fab_mapping_legacy:
        for v in variants:
            code = fab_mapping_legacy.get(v)
            if code:
                return {"raw": code, "exact": code, "prefix": None}
    return None


def _match_one_v1_adressen(
    row: CanonicalRow,
    db_v1: dict,
    adressen: dict | None,
    fab_mapping_legacy: dict[str, str] | None,
) -> MatchResult:
    """v1 (legacy 86k) cascade with Adressen-driven fab resolution + wildcards.

    Differs from ``_match_one``:
      - Resolves fab via Adressen first (738 entries with wildcards),
        falls back to legacy 24-entry mapping.
      - Supports ``SIE%``-style prefix wildcards: tries every fab-code in
        v1.by_type_to_fabs[type] that starts with the prefix.
    """
    candidates = _candidate_typenrs(row)
    if not candidates:
        return MatchResult(status="GEEN TYPE NR")

    fab_info = _resolve_fab_v1_adressen(row.manufacturer, adressen, fab_mapping_legacy)

    by_fab_type     = db_v1.get("by_fab_type") or {}
    by_type_to_fabs = db_v1.get("by_type_to_fabs") or {}
    by_type_only    = db_v1.get("by_type_only") or {}

    last_norm = ""
    best_ambiguous: MatchResult | None = None

    for cand in candidates:
        norm = _norm_type(cand)
        if not norm:
            continue
        last_norm = norm

        if fab_info:
            # Collect hits across the (exact code OR all wildcard-matched codes)
            hits: list[dict] = []
            matched_fc = ""
            if fab_info.get("exact"):
                fc = fab_info["exact"]
                fc_hits = by_fab_type.get((fc, norm), [])
                if fc_hits:
                    hits.extend(fc_hits)
                    matched_fc = fc
            elif fab_info.get("prefix"):
                pfx = fab_info["prefix"]
                for fc, recs in (by_type_to_fabs.get(norm) or {}).items():
                    if fc.startswith(pfx):
                        hits.extend(recs)
                        matched_fc = matched_fc or fc
            if len(hits) == 1:
                h = hits[0]
                return MatchResult(
                    status="MATCH", procos_artikel=h["artikel"],
                    procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
                    mapped_fab=matched_fc, matched_typenr=cand, n_hits=1,
                    matched_route="v1:fab+type" + (":wild" if fab_info.get("prefix") else ""),
                )
            if len(hits) > 1 and best_ambiguous is None:
                h = hits[0]
                best_ambiguous = MatchResult(
                    status="NIET UNIEK", procos_artikel=h["artikel"],
                    procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
                    mapped_fab=matched_fc, matched_typenr=cand, n_hits=len(hits),
                    matched_route="v1:fab+type" + (":wild" if fab_info.get("prefix") else ""),
                )

        # Type-only fallback (when no fab info, OR fab info gave no hits).
        hits = by_type_only.get(norm, [])
        if len(hits) == 1:
            h = hits[0]
            return MatchResult(
                status="MATCH (op type alleen)", procos_artikel=h["artikel"],
                procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
                mapped_fab="", matched_typenr=cand, n_hits=1,
                matched_route="v1:type-only",
            )
        if len(hits) > 1 and best_ambiguous is None:
            h = hits[0]
            best_ambiguous = MatchResult(
                status="NIET UNIEK (op type alleen)", procos_artikel=h["artikel"],
                procos_fabcode=h["fabrikant"], procos_omschrijving=h["omschrijving"],
                mapped_fab="", matched_typenr=cand, n_hits=len(hits),
                matched_route="v1:type-only",
            )

    if best_ambiguous is not None:
        return best_ambiguous
    return MatchResult(
        status="NIET GEVONDEN" if fab_info else "NIET GEVONDEN (fab niet gemapt)",
        mapped_fab=(fab_info or {}).get("raw", ""),
        matched_typenr=last_norm,
    )


def match_rows_combined(
    rows: Iterable[CanonicalRow],
    db_v2: dict | None,
    db_v1: dict | None,
    fab_mapping_v1: dict[str, str],
    klant_db: dict | None = None,
    klant_code: str | None = None,
    import_refs: dict | None = None,
) -> list[MatchResult]:
    """Match each row against v2 first; fall back to v1 on miss.

    Optional inputs:
      - klant_db + klant_code : enable cascade step 0 (klant-ref lookup).
      - import_refs           : enable Adressen-driven v1 fab-mapping
                                 with wildcards (replaces the legacy
                                 24-entry hardcoded fab_mapping).

    Either DB may be None — the cascade tries whatever it has. If both
    are None, every row returns NIET GEVONDEN.
    """
    adressen = (import_refs or {}).get("adressen") or None

    results: list[MatchResult] = []
    for row in rows:
        result_v2: MatchResult | None = None
        if db_v2:
            result_v2 = _match_one_v2(row, db_v2, klant_db, klant_code)
            if result_v2.status.startswith("MATCH"):
                results.append(result_v2)
                continue
        # Fallback to v1 when v2 missed (or no v2 DB available).
        if db_v1:
            if adressen:
                # Adressen-aware path (Fase C): richer 738-entry fab map
                # plus wildcard support.
                result_v1 = _match_one_v1_adressen(
                    row, db_v1, adressen, fab_mapping_v1 or {},
                )
            else:
                # Legacy path (Fase A/B): hardcoded 24-entry fab map only.
                result_v1 = _match_one(
                    row,
                    db_v1.get("by_fab_type") or {},
                    db_v1.get("by_type_only") or {},
                    fab_mapping_v1 or {},
                )
            if result_v1.status.startswith("MATCH"):
                if not result_v1.matched_route:
                    result_v1.matched_route = "v1:fab+type"
                results.append(result_v1)
                continue
            # Both missed — prefer the more informative v2 result if we had one,
            # otherwise fall back to v1's no-match result.
            results.append(result_v2 if result_v2 is not None else result_v1)
            continue
        # No v1 fallback configured: use whatever v2 produced (or generic miss).
        results.append(result_v2 if result_v2 is not None else MatchResult(status="GEEN TYPE NR"))
    return results


def summarize(results: list[MatchResult]) -> dict:
    """Aggregate stats over a list of MatchResult.

    Returns:
        {
            "total":       int,
            "by_status":   {status_string: count, ...},
            "match_count": int (any MATCH variant),
            "match_pct":   float (0..100, 1 decimal)
        }
    """
    total = len(results)
    by_status: dict[str, int] = defaultdict(int)
    match_count = 0
    for r in results:
        by_status[r.status] += 1
        if r.status.startswith("MATCH"):
            match_count += 1

    pct = (100.0 * match_count / total) if total else 0.0
    return {
        "total": total,
        "by_status": dict(by_status),
        "match_count": match_count,
        "match_pct": round(pct, 1),
    }
